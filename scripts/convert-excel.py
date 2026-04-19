#!/usr/bin/env python3
"""
Convert Robin's and Shino's old-format Excel files into the canonical
import template used by the fi-nance app.

Template columns:
  Date | Member | Category | Account | Symbol | Currency | Amount | UnitPriceJPY

Usage:
  python3 scripts/convert-excel.py
  Reads from ../../Financials-review.xlsx and ../../Shino-financials-review.xlsx
  Writes to data/import-template.xlsx
"""

import os, sys, re
from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
CONTAINER_DIR = os.path.dirname(os.path.dirname(PROJECT_DIR))

ROBIN_FILE = os.path.join(CONTAINER_DIR, "Financials-review.xlsx")
SHINO_FILE = os.path.join(CONTAINER_DIR, "Shino-financials-review.xlsx")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "data", "import-template.xlsx")

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

def val(cell):
    """Get cell value, returning None for empty."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "" or v == ".":
            return None
    return v

def num(cell):
    """Get numeric value from cell."""
    v = val(cell)
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0

def find_month_blocks(ws, max_row):
    """Find all month header rows. Returns list of (row, month_name, month_num)."""
    blocks = []
    for r in range(1, max_row + 1):
        for c in range(1, 20):
            v = val(ws.cell(r, c))
            if isinstance(v, str) and v.lower() in MONTHS:
                blocks.append((r, v, MONTHS[v.lower()]))
                break
    return blocks


def parse_robin_block(ws, header_row, year):
    """Parse one monthly block from Robin's sheet.

    Layout (relative to header_row):
      +0: Month name (col C)
      +1: "Robin" | ... | "Precious metals" | ... | "Price 1 OZ" | ... | "Crypto" | ... | "Price BTC"
      +2: Savings | val | gold_name | gold_qty | gold_price | gold_value | Alts | alts_display | btc_price | alts_value | Bank | bank_total
      +3: Stocks | val | ... | ... | ... | ... | Trading Acc | trading_qty | ... | trading_value | stocks | stocks_total
      +4: (maybe Cash Yen) or blank | ... | ... | ... | Cold | cold_qty | ... | cold_value | Metals | metals_total
      ...
      Summary rows: EURJPY | rate
    """
    rows = []
    month_name = val(ws.cell(header_row, 3))
    if not month_name:
        return rows
    month_num = MONTHS.get(month_name.lower())
    if not month_num:
        return rows
    date_str = f"{year}-{month_num:02d}-24"

    # Find the data start row (row after header with "Robin")
    data_start = header_row + 2  # First data row

    # Read bank/cash accounts by scanning column C labels (row order varies)
    savings_val = 0.0
    stocks_val = 0.0
    cash_yen = 0.0
    bank_yen = 0.0
    for offset in range(0, 6):
        label = val(ws.cell(data_start + offset, 3))
        if not isinstance(label, str):
            continue
        ll = label.lower().strip()
        v = num(ws.cell(data_start + offset, 4))
        if ll == "savings":
            savings_val = v
        elif ll == "stocks":
            stocks_val = v
        elif "cash" in ll and "yen" in ll:
            cash_yen = v
        elif "bank" in ll and "yen" in ll:
            bank_yen = v

    # Read gold: col F=qty, col G=price_eur, col H=value_eur
    gold_qty = num(ws.cell(data_start, 6))
    gold_price_eur = num(ws.cell(data_start, 7))

    # Read BTC price (EUR) from col K of first data row
    btc_price_eur = num(ws.cell(data_start, 11))

    # Read crypto values from col L (the VALUE column, which is authoritative)
    alts_value_eur = num(ws.cell(data_start, 12))
    trading_value_eur = num(ws.cell(data_start + 1, 12))
    cold_value_eur = num(ws.cell(data_start + 2, 12))

    # Find EURJPY rate from summary rows
    eurjpy = 0.0
    for scan in range(data_start, min(data_start + 12, ws.max_row + 1)):
        label = val(ws.cell(scan, 13))
        if isinstance(label, str) and "eurjpy" in label.lower():
            eurjpy = num(ws.cell(scan, 14))
            break

    if eurjpy <= 0:
        eurjpy = 160.0  # fallback

    # Convert EUR prices to JPY
    gold_price_jpy = gold_price_eur * eurjpy
    btc_price_jpy = btc_price_eur * eurjpy

    # Derive actual BTC quantities from value / price
    alts_btc = alts_value_eur / btc_price_eur if btc_price_eur > 0 else 0
    trading_btc = trading_value_eur / btc_price_eur if btc_price_eur > 0 else 0
    cold_btc = cold_value_eur / btc_price_eur if btc_price_eur > 0 else 0

    # Bank accounts (EUR)
    if savings_val != 0 or True:  # always include
        rows.append([date_str, "Robin", "bank", "Savings", None, "EUR", savings_val, eurjpy])
    if stocks_val != 0 or True:
        rows.append([date_str, "Robin", "bank", "Stocks", None, "EUR", stocks_val, eurjpy])

    # Cash Yen (JPY)
    if cash_yen != 0:
        rows.append([date_str, "Robin", "cash", "Cash Yen", None, "JPY", cash_yen, 1])

    # Bank Yen (JPY)
    if bank_yen != 0:
        rows.append([date_str, "Robin", "bank", "Bank Yen", None, "JPY", bank_yen, 1])

    # Gold
    rows.append([date_str, "Robin", "precious_metal", "Gold Coin", "XAU", "JPY", gold_qty, gold_price_jpy])

    # Crypto — always emit rows so zero transitions create removal entries
    rows.append([date_str, "Robin", "crypto", "BTC Alts", "bitcoin", "JPY", alts_btc, btc_price_jpy])
    rows.append([date_str, "Robin", "crypto", "BTC Trading", "bitcoin", "JPY", trading_btc, btc_price_jpy])
    rows.append([date_str, "Robin", "crypto", "BTC Cold", "bitcoin", "JPY", cold_btc, btc_price_jpy])

    return rows


def parse_shino_block(ws, header_row, year):
    """Parse one monthly block from Shino's sheet.

    Layout (relative to header_row):
      +0: Month name (col C)
      +1: "Shino" | ... | "Precious metals" | ... | "Price 1 OZ" | ... | "Crypto" | ... | "Price BTC"
      +2: Yen | (blank) | gold_name | gold_qty | gold_price_jpy | gold_value | Alts | 0 | btc_price_jpy | 0 | Cash | cash_val
      +3: ... | ... | ... | Trading Acc | 0 | ...
      +4: ... | ... | ... | Cold | cold_display | ... | cold_value | Metals | metals_total
    """
    rows = []
    month_name = val(ws.cell(header_row, 3))
    if not month_name:
        return rows
    month_num = MONTHS.get(month_name.lower())
    if not month_num:
        return rows
    date_str = f"{year}-{month_num:02d}-24"

    data_start = header_row + 2

    # Cash from col N (summary area) on first data row
    cash_val = num(ws.cell(data_start, 14))
    # Also check col D for "Yen" row value
    # Sometimes the cash is in col N with label in col M

    # Check for Cash label in col M
    cash_label = val(ws.cell(data_start, 13))
    if cash_label and "cash" in str(cash_label).lower():
        cash_val = num(ws.cell(data_start, 14))

    # Gold: col F=qty, col G=price_jpy, col H=value_jpy
    gold_qty = num(ws.cell(data_start, 6))
    gold_price_jpy = num(ws.cell(data_start, 7))

    # BTC price JPY from col K first data row
    btc_price_jpy = num(ws.cell(data_start, 11))

    # Cold wallet value from col L
    # Alts is always 0 for Shino, Trading Acc always 0
    # Cold is at data_start + 2
    cold_value_jpy = num(ws.cell(data_start + 2, 12))

    # Derive BTC from value
    cold_btc = cold_value_jpy / btc_price_jpy if btc_price_jpy > 0 else 0

    # Cash (JPY)
    if cash_val != 0:
        rows.append([date_str, "Shino", "cash", "Cash", None, "JPY", cash_val, 1])

    # Gold
    if gold_qty != 0:
        rows.append([date_str, "Shino", "precious_metal", "Gold Coin", "XAU", "JPY", gold_qty, gold_price_jpy])

    # Crypto (Cold only, Alts/Trading always 0)
    if cold_btc != 0:
        rows.append([date_str, "Shino", "crypto", "BTC Cold", "bitcoin", "JPY", cold_btc, btc_price_jpy])

    return rows


def process_robin():
    """Parse all sheets from Robin's workbook."""
    wb = load_workbook(ROBIN_FILE, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        blocks = find_month_blocks(ws, ws.max_row)
        for row_num, month_name, month_num in blocks:
            parsed = parse_robin_block(ws, row_num, year)
            all_rows.extend(parsed)
            if parsed:
                print(f"  Robin {year}-{month_num:02d} ({month_name}): {len(parsed)} rows")

    wb.close()
    return all_rows


def process_shino():
    """Parse all sheets from Shino's workbook."""
    wb = load_workbook(SHINO_FILE, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        blocks = find_month_blocks(ws, ws.max_row)
        for row_num, month_name, month_num in blocks:
            parsed = parse_shino_block(ws, row_num, year)
            all_rows.extend(parsed)
            if parsed:
                print(f"  Shino {year}-{month_num:02d} ({month_name}): {len(parsed)} rows")

    wb.close()
    return all_rows


def write_template(all_rows, output_path):
    """Write rows to the canonical template XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshots"

    headers = ["Date", "Member", "Category", "Account", "Symbol", "Currency", "Amount", "UnitPriceJPY"]
    ws.append(headers)

    # Sort by date, then member, then category
    cat_order = {"bank": 0, "cash": 1, "precious_metal": 2, "crypto": 3}
    all_rows.sort(key=lambda r: (r[0], r[1], cat_order.get(r[2], 99), r[3]))

    for row in all_rows:
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    print(f"\nWrote {len(all_rows)} rows to {output_path}")


def main():
    print(f"Robin file: {ROBIN_FILE}")
    print(f"Shino file: {SHINO_FILE}")

    if not os.path.exists(ROBIN_FILE):
        print(f"ERROR: {ROBIN_FILE} not found")
        sys.exit(1)
    if not os.path.exists(SHINO_FILE):
        print(f"ERROR: {SHINO_FILE} not found")
        sys.exit(1)

    print("\nParsing Robin's data...")
    robin_rows = process_robin()

    print("\nParsing Shino's data...")
    shino_rows = process_shino()

    all_rows = robin_rows + shino_rows
    write_template(all_rows, OUTPUT_FILE)

    # Print summary
    print("\n=== Summary ===")
    members = set(r[1] for r in all_rows)
    for member in sorted(members):
        member_rows = [r for r in all_rows if r[1] == member]
        dates = sorted(set(r[0] for r in member_rows))
        accounts = sorted(set(r[3] for r in member_rows))
        print(f"\n{member}: {len(dates)} months, {len(accounts)} accounts")
        print(f"  Date range: {dates[0]} to {dates[-1]}")
        print(f"  Accounts: {', '.join(accounts)}")

        # Print first and last month details
        for d in [dates[0], dates[-1]]:
            print(f"\n  {d}:")
            for r in member_rows:
                if r[0] == d:
                    amt = r[6]
                    price = r[7]
                    if r[2] in ("crypto", "precious_metal"):
                        print(f"    {r[3]}: qty={amt:.6f}, price_jpy={price:.0f}")
                    else:
                        print(f"    {r[3]}: {amt:.2f} {r[5]}")


if __name__ == "__main__":
    main()
